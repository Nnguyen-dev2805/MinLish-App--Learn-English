from __future__ import annotations

from io import BytesIO

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.deck import Deck
from app.models.user import User
from app.models.vocabulary_item import VocabularyItem
from app.services.anki_import_service import SOURCE_NAME


def test_deck_endpoints_require_auth(client: TestClient) -> None:
    response = client.get("/api/v1/decks")

    assert response.status_code == 401
    assert response.json()["code"] == "UNAUTHORIZED"


def test_list_seed_decks_and_unit_items(client: TestClient) -> None:
    _seed_anki_decks(client)
    headers = _auth_headers(client)

    decks_response = client.get("/api/v1/decks", headers=headers)

    assert decks_response.status_code == 200
    decks = decks_response.json()["items"]
    assert len(decks) == 30
    unit_01 = next(deck for deck in decks if deck["source_unit"] == "Unit 01")
    assert unit_01["name"] == "Unit 01"
    assert unit_01["is_public"] is True
    assert unit_01["is_seed"] is True
    assert unit_01["is_read_only"] is True
    assert unit_01["word_count"] == 20
    assert unit_01["learned_count"] == 0

    items_response = client.get(f"/api/v1/decks/{unit_01['id']}/items", headers=headers)

    assert items_response.status_code == 200
    items = items_response.json()["items"]
    assert len(items) == 20
    anxious = next(item for item in items if item["word"] == "anxious")
    assert anxious["meaning"] == "lo âu, băn khoăn"
    assert anxious["pronunciation"] == "['æŋ(k)ʃəs]"
    assert anxious["image_url"] == "/static/media/anki/book2/4000B2_601.jpg"
    assert anxious["word_audio_url"] == "/static/media/anki/book2/4000B2_anxious.mp3"

    review_response = client.post(
        "/api/v1/learning/reviews",
        headers=headers,
        json={
            "vocabulary_item_id": anxious["id"],
            "rating": "Good",
            "response_ms": 1200,
        },
    )
    assert review_response.status_code == 200

    deck_detail_response = client.get(f"/api/v1/decks/{unit_01['id']}", headers=headers)
    assert deck_detail_response.status_code == 200
    assert deck_detail_response.json()["learned_count"] == 1


def test_user_deck_and_item_crud(client: TestClient) -> None:
    headers = _auth_headers(client)

    create_deck_response = client.post(
        "/api/v1/decks",
        headers=headers,
        json={
            "name": "My IELTS Words",
            "description": "Personal deck",
            "tags": ["ielts", "personal"],
        },
    )

    assert create_deck_response.status_code == 201
    deck = create_deck_response.json()
    assert deck["name"] == "My IELTS Words"
    assert deck["word_count"] == 0
    assert deck["learned_count"] == 0
    assert deck["is_public"] is False
    assert deck["is_seed"] is False
    assert deck["is_read_only"] is False

    deck_id = deck["id"]
    update_deck_response = client.patch(
        f"/api/v1/decks/{deck_id}",
        headers=headers,
        json={"name": "My Travel Words", "tags": ["travel"]},
    )
    assert update_deck_response.status_code == 200
    assert update_deck_response.json()["name"] == "My Travel Words"
    assert update_deck_response.json()["tags"] == ["travel"]

    create_item_response = client.post(
        f"/api/v1/decks/{deck_id}/items",
        headers=headers,
        json={
            "word": "boarding pass",
            "pronunciation": None,
            "meaning": "thẻ lên máy bay",
            "description": "A document that lets a passenger board a plane.",
            "example": "Please show your boarding pass.",
            "collocation": "print a boarding pass",
            "related_words": ["airport"],
            "note": "travel vocabulary",
        },
    )
    assert create_item_response.status_code == 201
    item = create_item_response.json()
    assert item["word"] == "boarding pass"
    assert item["meaning"] == "thẻ lên máy bay"
    assert item["image_url"] is None

    item_id = item["id"]
    update_item_response = client.patch(
        f"/api/v1/items/{item_id}",
        headers=headers,
        json={"meaning": "vé lên máy bay", "related_words": ["airport", "gate"]},
    )
    assert update_item_response.status_code == 200
    assert update_item_response.json()["meaning"] == "vé lên máy bay"
    assert update_item_response.json()["related_words"] == ["airport", "gate"]

    deck_detail_response = client.get(f"/api/v1/decks/{deck_id}", headers=headers)
    assert deck_detail_response.status_code == 200
    assert deck_detail_response.json()["word_count"] == 1

    delete_item_response = client.delete(f"/api/v1/items/{item_id}", headers=headers)
    assert delete_item_response.status_code == 204

    delete_deck_response = client.delete(f"/api/v1/decks/{deck_id}", headers=headers)
    assert delete_deck_response.status_code == 204

    missing_deck_response = client.get(f"/api/v1/decks/{deck_id}", headers=headers)
    assert missing_deck_response.status_code == 404


def test_seed_deck_is_read_only(client: TestClient) -> None:
    seed_ids = _seed_anki_decks(client)
    headers = _auth_headers(client)
    seed_deck_id = seed_ids["Unit 01"]

    patch_response = client.patch(
        f"/api/v1/decks/{seed_deck_id}",
        headers=headers,
        json={"name": "Changed"},
    )
    assert patch_response.status_code == 403
    assert patch_response.json()["code"] == "SEED_DECK_READ_ONLY"

    delete_response = client.delete(f"/api/v1/decks/{seed_deck_id}", headers=headers)
    assert delete_response.status_code == 403
    assert delete_response.json()["code"] == "SEED_DECK_READ_ONLY"

    create_item_response = client.post(
        f"/api/v1/decks/{seed_deck_id}/items",
        headers=headers,
        json={"word": "blocked", "meaning": "không được thêm"},
    )
    assert create_item_response.status_code == 403
    assert create_item_response.json()["code"] == "SEED_DECK_READ_ONLY"

    export_response = client.get(f"/api/v1/decks/{seed_deck_id}/export", headers=headers)
    assert export_response.status_code == 404
    assert export_response.json()["code"] == "PERSONAL_DECK_NOT_FOUND"


def test_import_excel_only_for_personal_deck(client: TestClient) -> None:
    headers = _auth_headers(client)
    deck_response = client.post(
        "/api/v1/decks",
        headers=headers,
        json={"name": "Personal Excel Deck", "description": None, "tags": []},
    )
    assert deck_response.status_code == 201
    deck_id = deck_response.json()["id"]

    excel_bytes = _excel_bytes(
        [
            ["Word", "Pronunciation", "Meaning", "English Description", "Example"],
            ["hello", "/həˈloʊ/", "xin chào", "A greeting.", "Hello, nice to meet you."],
            ["", "", "missing word", "skip", "skip"],
            ["travel", None, "du lịch", None, "I love to travel."],
            ["broken", None, None, "missing meaning", "skip"],
        ],
    )
    import_response = client.post(
        f"/api/v1/decks/{deck_id}/import",
        headers=headers,
        files={
            "file": (
                "words.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert import_response.status_code == 200
    assert import_response.json()["imported_count"] == 2

    items_response = client.get(f"/api/v1/decks/{deck_id}/items", headers=headers)
    assert items_response.status_code == 200
    items = items_response.json()["items"]
    assert [item["word"] for item in items] == ["hello", "travel"]
    assert items[0]["pronunciation"] == "/həˈloʊ/"
    assert items[0]["meaning"] == "xin chào"
    assert items[0]["description"] == "A greeting."
    assert items[0]["example"] == "Hello, nice to meet you."
    assert items[0]["collocation"] is None
    assert items[0]["related_words"] is None
    assert items[0]["note"] is None


def test_export_excel_only_for_personal_deck(client: TestClient) -> None:
    headers = _auth_headers(client)
    deck_response = client.post(
        "/api/v1/decks",
        headers=headers,
        json={"name": "Export Deck", "description": None, "tags": []},
    )
    assert deck_response.status_code == 201
    deck_id = deck_response.json()["id"]

    create_item_response = client.post(
        f"/api/v1/decks/{deck_id}/items",
        headers=headers,
        json={
            "word": "airport",
            "pronunciation": "/ˈeəpɔːt/",
            "meaning": "sân bay",
            "description": "A place where planes land and take off.",
            "example": "We arrived at the airport early.",
        },
    )
    assert create_item_response.status_code == 201

    export_response = client.get(f"/api/v1/decks/{deck_id}/export", headers=headers)

    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    workbook = openpyxl.load_workbook(BytesIO(export_response.content), data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == ("Word", "Pronunciation", "Meaning", "English Description", "Example")
    assert rows[1] == (
        "airport",
        "/ˈeəpɔːt/",
        "sân bay",
        "A place where planes land and take off.",
        "We arrived at the airport early.",
    )


def test_user_cannot_access_other_users_private_deck(client: TestClient) -> None:
    owner_headers = _auth_headers(client, email="owner@example.com")
    other_headers = _auth_headers(client, email="other@example.com")

    create_response = client.post(
        "/api/v1/decks",
        headers=owner_headers,
        json={"name": "Private deck", "description": None, "tags": []},
    )
    assert create_response.status_code == 201
    deck_id = create_response.json()["id"]

    get_response = client.get(f"/api/v1/decks/{deck_id}", headers=other_headers)
    assert get_response.status_code == 404

    patch_response = client.patch(
        f"/api/v1/decks/{deck_id}",
        headers=other_headers,
        json={"name": "Not mine"},
    )
    assert patch_response.status_code == 404


def _auth_headers(client: TestClient, email: str = "learner@example.com") -> dict[str, str]:
    response = client.post(
        "/api/v1/auth/register",
        json={
            "email": email,
            "password": "secret123",
            "name": "Min Learner",
        },
    )
    assert response.status_code == 200
    _verify_user_in_db(client, email)
    access_token = response.json()["access_token"]
    return {"Authorization": f"Bearer {access_token}"}


def _verify_user_in_db(client: TestClient, email: str) -> None:
    session_local = client.app.state.testing_session_local
    with session_local() as db:
        user = db.scalar(select(User).where(User.email == email))
        assert user is not None
        user.email_verified = True
        db.commit()


def _seed_anki_decks(client: TestClient) -> dict[str, int]:
    session_local = client.app.state.testing_session_local
    with session_local() as db:
        return _seed_anki_decks_in_session(db)


def _seed_anki_decks_in_session(db: Session) -> dict[str, int]:
    seed_ids: dict[str, int] = {}
    for unit_number in range(1, 31):
        source_unit = f"Unit {unit_number:02d}"
        deck = Deck(
            user_id=None,
            name=source_unit,
            description=SOURCE_NAME,
            tags=["4000-essential", "book-2", f"unit-{unit_number:02d}", "seed"],
            is_public=True,
            is_seed=True,
            is_read_only=True,
            source_name=SOURCE_NAME,
            source_unit=source_unit,
        )
        db.add(deck)
        db.flush()
        seed_ids[source_unit] = deck.id

        for index in range(1, 21):
            anki_number = 600 + ((unit_number - 1) * 20) + index
            if unit_number == 1 and index == 1:
                db.add(_anxious_item(deck.id))
            else:
                db.add(
                    VocabularyItem(
                        deck_id=deck.id,
                        word=f"word-{anki_number}",
                        meaning=f"meaning-{anki_number}",
                        anki_number=str(anki_number),
                        source_key=f"book2-{anki_number}",
                        source_name=SOURCE_NAME,
                        source_unit=source_unit,
                    ),
                )
        db.flush()
    db.commit()
    return seed_ids


def _excel_bytes(rows: list[list[object | None]]) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _anxious_item(deck_id: int) -> VocabularyItem:
    return VocabularyItem(
        deck_id=deck_id,
        word="anxious",
        pronunciation="['æŋ(k)ʃəs]",
        meaning="lo âu, băn khoăn",
        description="When a person is anxious, they worry that something bad will happen.",
        example="She was anxious about not making her appointment on time.",
        note="tính từ",
        suggestion="a__x__ __ __ __",
        anki_number="601",
        source_key="book2-601",
        source_name=SOURCE_NAME,
        source_unit="Unit 01",
        image_url="/static/media/anki/book2/4000B2_601.jpg",
        word_audio_url="/static/media/anki/book2/4000B2_anxious.mp3",
        meaning_audio_url="/static/media/anki/book2/4000B2_anxious_meaning.mp3",
        example_audio_url="/static/media/anki/book2/4000B2_anxious_example.mp3",
    )
