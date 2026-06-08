from io import BytesIO

import openpyxl
from fastapi import UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.exceptions import ApiError
from app.models.deck import Deck
from app.models.user import User
from app.models.vocabulary_item import VocabularyItem


class ImportExportService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def import_excel(self, user: User, deck: Deck, file: UploadFile) -> int:
        if not file.filename or not file.filename.endswith(".xlsx"):
            raise ApiError(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Chỉ hỗ trợ file định dạng .xlsx (Excel).",
                code="INVALID_FILE_FORMAT",
            )
        
        try:
            content = file.file.read()
            workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
            sheet = workbook.active
            
            imported_count = 0
            
            # Expected columns:
            # A: Word, B: Pronunciation, C: Meaning, D: English Description, E: Example
            for row in sheet.iter_rows(min_row=2, values_only=True):
                word = self._cell_text(row, 0)
                meaning = self._cell_text(row, 2)
                if not word or not meaning:
                    continue

                pronunciation = self._cell_text(row, 1)
                description = self._cell_text(row, 3)
                example = self._cell_text(row, 4)
                
                vocab_item = VocabularyItem(
                    deck_id=deck.id,
                    word=word,
                    pronunciation=pronunciation,
                    meaning=meaning,
                    description=description,
                    example=example,
                )
                self.db.add(vocab_item)
                imported_count += 1
                
            self.db.commit()
            return imported_count
            
        except Exception as e:
            self.db.rollback()
            raise ApiError(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Lỗi khi xử lý file Excel: {str(e)}",
                code="EXCEL_PROCESSING_ERROR",
            )

    def export_excel(self, deck: Deck) -> BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Vocabulary"
        sheet.append(["Word", "Pronunciation", "Meaning", "English Description", "Example"])

        items = self.db.scalars(
            select(VocabularyItem)
            .where(VocabularyItem.deck_id == deck.id)
            .order_by(VocabularyItem.word, VocabularyItem.id),
        ).all() # .all : lấy kết quả in ra thành list

        for item in items:
            sheet.append(
                [
                    item.word,
                    item.pronunciation or "",
                    item.meaning,
                    item.description or "",
                    item.example or "",
                ],
            )

        output = BytesIO()
        workbook.save(output)
        output.seek(0) # để con trỏ chuyền về đầu trang để cho frontend đọc
        return output

    @staticmethod
    def _cell_text(row: tuple[object, ...], index: int) -> str | None:
        if len(row) <= index or row[index] is None:
            return None
        value = str(row[index]).strip()
        return value or None
